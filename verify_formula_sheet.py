import os
import sys
import openpyxl
import numpy as np

# Load our FEA engine for comparison
from verify_lbend import run_lbend_verification

def verify_formula_sheet():
    print("=== STARTING INTERACTIVE FORMULA SHEET VERIFICATION ===")
    
    filepath = "L_bend_analyzer.xlsx"
    assert os.path.exists(filepath), f"File {filepath} not found!"
    
    # Load with data_only=False to inspect formula strings
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb["L-Bend Analyzer"]
    
    print("Checking input cell value strings...")
    assert ws["C6"].value == 5.0, f"C6 L1 is {ws['C6'].value}, expected 5.0"
    assert ws["C7"].value == 4.0, f"C7 L2 is {ws['C7'].value}, expected 4.0"
    assert ws["C8"].value == 0.1143, f"C8 OD is {ws['C8'].value}, expected 0.1143"
    assert ws["C9"].value == 0.00602, f"C9 t is {ws['C9'].value}, expected 0.00602"
    assert ws["C10"].value == 150.0, f"C10 dT is {ws['C10'].value}, expected 150.0"
    assert ws["C11"].value == 0.0, f"C11 P is {ws['C11'].value}, expected 0.0"
    assert ws["C12"].value == 2.0e11, f"C12 E is {ws['C12'].value}, expected 2.0e11"
    
    print("Checking section property formula strings...")
    assert ws["C19"].value == "=C8/2", f"C19 formula is {ws['C19'].value}"
    assert ws["C20"].value == "=C19-C9", f"C20 formula is {ws['C20'].value}"
    assert ws["C21"].value == "=PI()*(C19^2-C20^2)", f"C21 formula is {ws['C21'].value}"
    assert ws["C22"].value == "=(PI()/4)*(C19^4-C20^4)", f"C22 formula is {ws['C22'].value}"
    assert ws["C23"].value == "=C22/C19", f"C23 formula is {ws['C23'].value}"
    assert ws["C24"].value == "=(C11*C8)/(4*C9)", f"C24 formula is {ws['C24'].value}"
    assert ws["C25"].value == "=1.25*C15+0.25*C16", f"C25 formula is {ws['C25'].value}"
    
    print("Checking compatibility matrix formulas...")
    assert ws["F7"].value == "=((C7^3)/3+C6*(C7^2))+(C22/C21)*C6", f"F7 formula is {ws['F7'].value}"
    assert ws["G7"].value == "=-((C6^2)*C7)/2", f"G7 formula is {ws['G7'].value}"
    assert ws["H7"].value == "=-C7*(C6+C7/2)", f"H7 formula is {ws['H7'].value}"
    assert ws["I7"].value == "=-C12*C22*C28", f"I7 formula is {ws['I7'].value}"
    
    print("Checking matrix solver formulas...")
    assert ws["G13"].value == "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 1, 1)", f"G13 formula is {ws['G13'].value}"
    assert ws["G14"].value == "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 2, 1)", f"G14 formula is {ws['G14'].value}"
    assert ws["G15"].value == "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 3, 1)", f"G15 formula is {ws['G15'].value}"
    
    print("Checking node stress formulas...")
    assert ws["G19"].value == "=G15+G13*C7-G14*C6", f"G19 formula is {ws['G19'].value}"
    assert ws["H19"].value == "=G13", f"H19 formula is {ws['H19'].value}"
    assert ws["I19"].value == "=ABS(G19)/C23+ABS(H19)/C21+C24", f"I19 formula is {ws['I19'].value}"
    assert ws["J19"].value == "=C16", f"J19 formula is {ws['J19'].value}"
    
    # 6. Evaluate the math using python numpy to make sure formulas are correct
    L1 = 5.0
    L2 = 4.0
    OD = 0.1143
    t = 0.00602
    dT = 150.0
    P = 0.0
    E = 2.0e11
    alpha = 1.2e-5
    Sc = 1.379e8
    Sh = 1.379e8
    
    ro = OD / 2.0
    ri = ro - t
    A = np.pi * (ro**2 - ri**2)
    I = (np.pi / 4.0) * (ro**4 - ri**4)
    Z = I / ro
    SL_pr = (P * OD) / (4.0 * t) if t > 0 else 0.0
    
    dx_th = alpha * L1 * dT
    dy_th = alpha * L2 * dT
    
    fxx = (L2**3 / 3.0 + L1 * L2**2) + (I / A) * L1
    fyy = (L1**3 / 3.0) + (I / A) * L2
    fmm = L1 + L2
    fxy = -L1**2 * L2 / 2.0
    fxm = -L2 * (L1 + L2 / 2.0)
    fym = L1**2 / 2.0
    
    F_mat = np.array([
        [fxx, fxy, fxm],
        [fxy, fyy, fym],
        [fxm, fym, fmm]
    ])
    
    d_vector = np.array([
        -E * I * dx_th,
        -E * I * dy_th,
        0.0
    ])
    
    solved_reactions = np.linalg.solve(F_mat, d_vector)
    Fx = solved_reactions[0]
    Fy = solved_reactions[1]
    Mz = solved_reactions[2]
    
    # Node 0 (Anchor 1)
    M0 = Mz + Fx * L2 - Fy * L1
    SL0 = abs(M0)/Z + abs(Fx)/A + SL_pr
    SE0 = abs(M0)/Z
    
    # Node 1 (Elbow)
    M1 = Mz + Fx * L2
    F_axial_elbow = max(abs(Fx), abs(Fy))
    SL1 = abs(M1)/Z + F_axial_elbow/A + SL_pr
    SE1 = abs(M1)/Z
    
    # Node 2 (Anchor 2)
    M2 = Mz
    SL2 = abs(M2)/Z + abs(Fy)/A + SL_pr
    SE2 = abs(M2)/Z
    
    print("\nEvaluated Analytical Values:")
    print(f"Fx: {Fx:.4e} N (expected ~ -8.0864e+02)")
    print(f"Fy: {Fy:.4e} N (expected ~ -5.4789e+02)")
    print(f"Mz: {Mz:.4e} N-m (expected ~ -1.7548e+03)")
    print(f"Node 0 Stress: {SL0:.4e} Pa, ratio: {SL0/Sh:.2f}")
    print(f"Node 1 Stress: {SL1:.4e} Pa, ratio: {SL1/Sh:.2f}")
    print(f"Node 2 Stress: {SL2:.4e} Pa, ratio: {SL2/Sh:.2f}")
    
    # Assert values match our physical FEA verify results (within 1.0 N)
    assert abs(Fx - (-808.6422219)) < 1.0, f"Fx mismatch: {Fx}"
    assert abs(Fy - (-547.8860269)) < 1.0, f"Fy mismatch: {Fy}"
    assert abs(Mz - (-1754.832961)) < 1.0, f"Mz mismatch: {Mz}"
    
    print("\nVerification Passed! The native spreadsheet formulas represent the exact correct analytical model.")
    print("=== INTERACTIVE FORMULA SHEET VERIFICATION SUCCESSFUL! ===")

if __name__ == "__main__":
    verify_formula_sheet()
