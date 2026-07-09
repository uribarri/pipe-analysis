import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_interactive_workbook(filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "L-Bend Analyzer"
    
    # Enable grid lines
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    else:
        ws.sheet_view.showGridLines = True
        
    # Styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_data_bold = Font(name="Segoe UI", size=10, bold=True)
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="006100")
    font_fail = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
    
    fill_section = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_input = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow for inputs
    fill_calc = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Soft grey for calcs
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='double', color='1F4E79')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # --- Title Block ---
    ws['B2'] = "ASME B31.3 Piping L-Bend Stress Compliance Analyzer"
    ws['B2'].font = font_title
    ws['B3'] = "Interactive formula-based solver. Edit values in yellow cells; output updates instantly."
    ws['B3'].font = font_subtitle
    
    # --- Section 1: User Inputs ---
    ws.merge_cells('B5:D5')
    ws['B5'] = "1. DESIGN INPUTS"
    ws['B5'].font = font_section
    ws['B5'].fill = fill_section
    ws['B5'].alignment = align_center
    
    inputs = [
        ("Horizontal Run Length (L1)", 5.0, "m"),
        ("Vertical Run Length (L2)", 4.0, "m"),
        ("Pipe Outer Diameter (OD)", 0.1143, "m"),
        ("Nominal Wall Thickness (t)", 0.00602, "m"),
        ("Temperature Change (dT)", 150.0, "°C"),
        ("Internal Design Pressure (P)", 0.0, "Pa"),
        ("Young's Modulus (E)", 2.0e11, "Pa"),
        ("Shear Modulus (G)", 7.7e10, "Pa"),
        ("Thermal Expansion Coeff (alpha)", 1.2e-5, "1/°C"),
        ("Cold Allowable Stress (Sc)", 1.379e8, "Pa"),
        ("Hot Allowable Stress (Sh)", 1.379e8, "Pa")
    ]
    
    for idx, (name, val, unit) in enumerate(inputs, 6):
        ws.cell(row=idx, column=2, value=name).font = font_data
        ws.cell(row=idx, column=2).alignment = align_left
        ws.cell(row=idx, column=2).border = thin_border
        
        cell_val = ws.cell(row=idx, column=3, value=val)
        cell_val.font = font_data_bold
        cell_val.alignment = align_right
        cell_val.fill = fill_input
        cell_val.border = thin_border
        
        # Apply formatting
        if name == "Pipe Outer Diameter (OD)" or name == "Nominal Wall Thickness (t)":
            cell_val.number_format = '0.0000'
        elif "Modulus" in name or "Allowable" in name or "Pressure" in name:
            cell_val.number_format = '0.00E+00'
        elif name == "Thermal Expansion Coeff (alpha)":
            cell_val.number_format = '0.00E+00'
        else:
            cell_val.number_format = '0.0'
            
        ws.cell(row=idx, column=4, value=unit).font = font_data
        ws.cell(row=idx, column=4).alignment = align_center
        ws.cell(row=idx, column=4).border = thin_border

    # --- Section 2: Calculated Section Properties ---
    ws.merge_cells('B18:D18')
    ws['B18'] = "2. SECTION PROPERTIES"
    ws['B18'].font = font_section
    ws['B18'].fill = fill_section
    ws['B18'].alignment = align_center
    
    props = [
        ("Outer Radius (ro)", "=C8/2", "m", "0.0000"),
        ("Inner Radius (ri)", "=C19-C9", "m", "0.0000"),
        ("Cross-sectional Area (A)", "=PI()*(C19^2-C20^2)", "m²", "0.00E+00"),
        ("Moment of Inertia (I)", "=(PI()/4)*(C19^4-C20^4)", "m⁴", "0.00E+00"),
        ("Section Modulus (Z)", "=C22/C19", "m³", "0.00E+00"),
        ("Pressure Long. Stress (SL_pr)", "=(C11*C8)/(4*C9)", "Pa", "0.00E+00"),
        ("Allowable Expansion Stress (SA)", "=1.25*C15+0.25*C16", "Pa", "0.00E+00")
    ]
    
    for idx, (name, formula, unit, num_format) in enumerate(props, 19):
        ws.cell(row=idx, column=2, value=name).font = font_data
        ws.cell(row=idx, column=2).alignment = align_left
        ws.cell(row=idx, column=2).border = thin_border
        
        cell_val = ws.cell(row=idx, column=3, value=formula)
        cell_val.font = font_data
        cell_val.alignment = align_right
        cell_val.fill = fill_calc
        cell_val.border = thin_border
        cell_val.number_format = num_format
        
        ws.cell(row=idx, column=4, value=unit).font = font_data
        ws.cell(row=idx, column=4).alignment = align_center
        ws.cell(row=idx, column=4).border = thin_border

    # --- Section 3: Thermal Deflection Vector ---
    ws.merge_cells('B27:D27')
    ws['B27'] = "3. THERMAL EXPANSION"
    ws['B27'].font = font_section
    ws['B27'].fill = fill_section
    ws['B27'].alignment = align_center
    
    thermals = [
        ("Free X Deflection (dx_th)", "=C14*C6*C10", "m", "0.0000"),
        ("Free Y Deflection (dy_th)", "=C14*C7*C10", "m", "0.0000")
    ]
    
    for idx, (name, formula, unit, num_format) in enumerate(thermals, 28):
        ws.cell(row=idx, column=2, value=name).font = font_data
        ws.cell(row=idx, column=2).alignment = align_left
        ws.cell(row=idx, column=2).border = thin_border
        
        cell_val = ws.cell(row=idx, column=3, value=formula)
        cell_val.font = font_data
        cell_val.alignment = align_right
        cell_val.fill = fill_calc
        cell_val.border = thin_border
        cell_val.number_format = num_format
        
        ws.cell(row=idx, column=4, value=unit).font = font_data
        ws.cell(row=idx, column=4).alignment = align_center
        ws.cell(row=idx, column=4).border = thin_border

    # --- Section 4: Compatibility Matrix ---
    ws.merge_cells('F5:I5')
    ws['F5'] = "4. FLEXIBILITY / COMPATIBILITY MATRIX"
    ws['F5'].font = font_section
    ws['F5'].fill = fill_section
    ws['F5'].alignment = align_center
    
    # Headers
    headers_matrix = ["Flexibility f_iX", "Flexibility f_iY", "Flexibility f_iM", "Load Vector d_i"]
    for c_idx, h in enumerate(headers_matrix, 6):
        cell = ws.cell(row=6, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Formulas for Flexibility elements
    # Row 7 (X compatibility): fxx, fxy, fxm, Load X
    ws.cell(row=7, column=6, value="=((C7^3)/3+C6*(C7^2))+(C22/C21)*C6").number_format = '0.00E+00' # fxx
    ws.cell(row=7, column=7, value="=-((C6^2)*C7)/2").number_format = '0.00E+00'                # fxy
    ws.cell(row=7, column=8, value="=-C7*(C6+C7/2)").number_format = '0.00E+00'                # fxm
    ws.cell(row=7, column=9, value="=-C12*C22*C28").number_format = '0.00E+00'                 # dx_load
    
    # Row 8 (Y compatibility): fyx, fyy, fym, Load Y
    ws.cell(row=8, column=6, value="=G7").number_format = '0.00E+00'                            # fyx = fxy
    ws.cell(row=8, column=7, value="=((C6^3)/3)+(C22/C21)*C7").number_format = '0.00E+00'      # fyy
    ws.cell(row=8, column=8, value="=(C6^2)/2").number_format = '0.00E+00'                     # fym
    ws.cell(row=8, column=9, value="=-C12*C22*C29").number_format = '0.00E+00'                 # dy_load
    
    # Row 9 (M compatibility): fmx, fmy, fmm, Load M
    ws.cell(row=9, column=6, value="=H7").number_format = '0.00E+00'                            # fmx = fxm
    ws.cell(row=9, column=7, value="=H8").number_format = '0.00E+00'                            # fmy = fym
    ws.cell(row=9, column=8, value="=C6+C7").number_format = '0.00E+00'                         # fmm
    ws.cell(row=9, column=9, value=0.0).number_format = '0.00E+00'                              # dm_load = 0
    
    for r in range(7, 10):
        for c in range(6, 10):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.alignment = align_right
            cell.fill = fill_calc
            cell.border = thin_border

    # --- Section 5: Solved Reaction Forces ---
    ws.merge_cells('F11:H11')
    ws['F11'] = "5. SOLVED ANCHOR 2 REACTIONS"
    ws['F11'].font = font_section
    ws['F11'].fill = fill_section
    ws['F11'].alignment = align_center
    
    ws.cell(row=12, column=6, value="Reaction Component").font = font_header
    ws.cell(row=12, column=6).fill = fill_header
    ws.cell(row=12, column=6).border = thin_border
    
    ws.cell(row=12, column=7, value="Solved Force/Moment").font = font_header
    ws.cell(row=12, column=7).fill = fill_header
    ws.cell(row=12, column=7).border = thin_border
    
    ws.cell(row=12, column=8, value="Unit").font = font_header
    ws.cell(row=12, column=8).fill = fill_header
    ws.cell(row=12, column=8).border = thin_border
    
    reactions = [
        ("Fx (Horizontal Reaction Force)", "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 1, 1)", "N"),
        ("Fy (Vertical Reaction Force)", "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 2, 1)", "N"),
        ("Mz (Bending Moment)", "=INDEX(MMULT(MINVERSE(F7:H9), I7:I9), 3, 1)", "N-m")
    ]
    
    for idx, (name, formula, unit) in enumerate(reactions, 13):
        ws.cell(row=idx, column=6, value=name).font = font_data
        ws.cell(row=idx, column=6).alignment = align_left
        ws.cell(row=idx, column=6).border = thin_border
        
        cell_val = ws.cell(row=idx, column=7, value=formula)
        cell_val.font = font_data_bold
        cell_val.alignment = align_right
        cell_val.fill = fill_calc
        cell_val.border = thin_border
        cell_val.number_format = '#,##0.00'
        
        ws.cell(row=idx, column=8, value=unit).font = font_data
        ws.cell(row=idx, column=8).alignment = align_center
        ws.cell(row=idx, column=8).border = thin_border

    # --- Section 6: Stress Compliance Report ---
    ws.merge_cells('F17:P17')
    ws['F17'] = "6. ASME B31.3 STRESS COMPLIANCE REPORT"
    ws['F17'].font = font_section
    ws['F17'].fill = fill_section
    ws['F17'].alignment = align_center
    
    headers_stress = [
        "Node", "Bending Moment (N-m)", "Axial Force (N)",
        "Sustained Stress (Pa)", "Sustained Allowable (Pa)", "Sustained Ratio",
        "Expansion Stress (Pa)", "Expansion Allowable (Pa)", "Expansion Ratio",
        "Max Ratio", "Status"
    ]
    for c_idx, h in enumerate(headers_stress, 6):
        cell = ws.cell(row=18, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Row 19: Node 0 (Anchor 1)
    ws.cell(row=19, column=6, value="0 (Anchor 1)").alignment = align_center
    ws.cell(row=19, column=7, value="=G15+G13*C7-G14*C6")   # M0 = Mz + Fx*L2 - Fy*L1
    ws.cell(row=19, column=8, value="=G13")                 # Fx
    ws.cell(row=19, column=9, value="=ABS(G19)/C23+ABS(H19)/C21+C24") # SL0 = |M0|/Z + |Fx|/A + SL_pr
    ws.cell(row=19, column=10, value="=C16")                # Sh
    ws.cell(row=19, column=11, value="=I19/J19")            # Sustained Ratio
    ws.cell(row=19, column=12, value="=ABS(G19)/C23")       # SE0 = |M0|/Z
    ws.cell(row=19, column=13, value="=C25")                # SA
    ws.cell(row=19, column=14, value="=L19/M19")            # Expansion Ratio
    ws.cell(row=19, column=15, value="=MAX(K19,N19)")       # Max Ratio
    ws.cell(row=19, column=16, value='=IF(O19<=1.0,"PASS","FAIL")').alignment = align_center
    
    # Row 20: Node 1 (Corner Elbow)
    ws.cell(row=20, column=6, value="1 (Corner Elbow)").alignment = align_center
    ws.cell(row=20, column=7, value="=G15+G13*C7")          # M1 = Mz + Fx*L2
    ws.cell(row=20, column=8, value="=MAX(ABS(G13),ABS(G14))") # Conservative axial force
    ws.cell(row=20, column=9, value="=ABS(G20)/C23+ABS(H20)/C21+C24") # SL1
    ws.cell(row=20, column=10, value="=C16")
    ws.cell(row=20, column=11, value="=I20/J20")
    ws.cell(row=20, column=12, value="=ABS(G20)/C23")
    ws.cell(row=20, column=13, value="=C25")
    ws.cell(row=20, column=14, value="=L20/M20")
    ws.cell(row=20, column=15, value="=MAX(K20,N20)")
    ws.cell(row=20, column=16, value='=IF(O20<=1.0,"PASS","FAIL")').alignment = align_center

    # Row 21: Node 2 (Anchor 2)
    ws.cell(row=21, column=6, value="2 (Anchor 2)").alignment = align_center
    ws.cell(row=21, column=7, value="=G15")                 # M2 = Mz
    ws.cell(row=21, column=8, value="=G14")                 # Fy
    ws.cell(row=21, column=9, value="=ABS(G21)/C23+ABS(H21)/C21+C24") # SL2
    ws.cell(row=21, column=10, value="=C16")
    ws.cell(row=21, column=11, value="=I21/J21")
    ws.cell(row=21, column=12, value="=ABS(G21)/C23")
    ws.cell(row=21, column=13, value="=C25")
    ws.cell(row=21, column=14, value="=L21/M21")
    ws.cell(row=21, column=15, value="=MAX(K21,N21)")
    ws.cell(row=21, column=16, value='=IF(O21<=1.0,"PASS","FAIL")').alignment = align_center

    # Apply data formatting and fonts to report table
    for r in range(19, 22):
        for c in range(6, 17):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c == 6:
                cell.font = font_data_bold
            elif c == 16:
                pass # Special font setting below
            else:
                cell.font = font_data
                cell.alignment = align_right
                
            # Formatting categories
            if c in (7, 8):
                cell.number_format = '#,##0.0'
            elif c in (9, 10, 12, 13):
                cell.number_format = '0.00E+00'
            elif c in (11, 14, 15):
                cell.number_format = '0.00'
                
            # Zebra striping
            if r % 2 == 1:
                cell.fill = fill_zebra
                
    # Format PASS/FAIL cells
    for r in range(19, 22):
        cell_status = ws.cell(row=r, column=16)
        cell_status.font = font_pass
        cell_status.fill = fill_pass
        
    # Add native Excel conditional formatting for PASS/FAIL
    from openpyxl.formatting.rule import CellIsRule
    rule_pass = CellIsRule(operator='equal', formula=['"PASS"'], stopIfTrue=True, font=font_pass, fill=fill_pass)
    rule_fail = CellIsRule(operator='equal', formula=['"FAIL"'], stopIfTrue=True, font=font_fail, fill=fill_fail)
    ws.conditional_formatting.add('P19:P21', rule_pass)
    ws.conditional_formatting.add('P19:P21', rule_fail)
    
    # Add conditional formatting for ratios > 1.0 (light red highlight)
    rule_ratio = CellIsRule(operator='greaterThan', formula=['1.0'], stopIfTrue=True, font=font_fail, fill=fill_fail)
    ws.conditional_formatting.add('K19:K21', rule_ratio)
    ws.conditional_formatting.add('N19:N21', rule_ratio)
    ws.conditional_formatting.add('O19:O21', rule_ratio)

    # Double bottom border on totals / end row
    for col in range(6, 17):
        ws.cell(row=21, column=col).border = double_bottom_border

    # Adjust columns width
    autofit_columns(ws)
    
    # Save workbook
    wb.save(filepath)
    print(f"Formula-based interactive workbook saved to {filepath}")

def autofit_columns(ws, padding=3, min_width=12):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):
                val = "123,456.00" # Dummy size for formulas
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

if __name__ == "__main__":
    filepath = "L_bend_analyzer.xlsx"
    generate_interactive_workbook(filepath)
