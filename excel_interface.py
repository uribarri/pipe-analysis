import os
import argparse
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Import our FEA engine
from fea_engine import FEAEngine

def parse_bc_value(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip().upper()
    if val_str in ('TRUE', 'YES', 'Y', 'FIXED', '1'):
        return True
    if val_str in ('FALSE', 'NO', 'N', 'FREE', '0', ''):
        return False
    try:
        return float(val)
    except ValueError:
        return False

def style_worksheet(ws):
    # Ensure grid lines are visible
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    else:
        ws.sheet_view.showGridLines = True

def apply_table_styles(ws, start_row, start_col, end_row, end_col, title_rows=1):
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    
    data_font = Font(name='Segoe UI', size=10)
    zebra_fill = PatternFill(start_color='F2F6F9', end_color='F2F6F9', fill_type='solid')
    
    # Headers
    for row in range(start_row, start_row + title_rows):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
    # Data rows
    for row in range(start_row + title_rows, end_row + 1):
        is_zebra = (row - start_row - title_rows) % 2 == 1
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if not cell.fill.fill_type: # Keep custom cell fills if already set
                if is_zebra:
                    cell.fill = zebra_fill
            # General alignment
            if col == start_col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

def autofit_columns(ws, padding=3, min_width=10):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            # Don't let long formula strings expand columns infinitely
            if val.startswith('='):
                val = "12345678"
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

def generate_template(filepath: str):
    """Generates an Excel template sheet with sample piping system input data."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Fonts and borders
    title_font = Font(name='Segoe UI', size=14, bold=True, color='1F4E79')
    
    # 1. Materials Sheet
    ws_mat = wb.create_sheet(title='Materials')
    style_worksheet(ws_mat)
    ws_mat.cell(row=1, column=1, value="Material Properties").font = title_font
    
    headers_mat = ["Material ID", "E (Pa)", "G (Pa)", "alpha (1/°C)", "Yield Strength (Pa)", "Sc (Pa)", "Sh (Pa)"]
    data_mat = [
        ["mat_id_1", 2.0e11, 7.7e10, 1.2e-5, 2.5e8, 1.379e8, 1.379e8]
    ]
    
    for c_idx, h in enumerate(headers_mat, 1):
        ws_mat.cell(row=3, column=c_idx, value=h)
    for r_idx, row in enumerate(data_mat, 4):
        for c_idx, val in enumerate(row, 1):
            ws_mat.cell(row=r_idx, column=c_idx, value=val)
            if c_idx > 1:
                ws_mat.cell(row=r_idx, column=c_idx).number_format = '0.00E+00'
                
    apply_table_styles(ws_mat, start_row=3, start_col=1, end_row=3+len(data_mat), end_col=len(headers_mat))
    autofit_columns(ws_mat)

    # 2. Sections Sheet
    ws_sec = wb.create_sheet(title='Sections')
    style_worksheet(ws_sec)
    ws_sec.cell(row=1, column=1, value="Section Geometries").font = title_font
    
    headers_sec = ["Section ID", "Outer Diameter (m)", "Wall Thickness (m)", "Type", "Fluid Density (kg/m³)", "Insulation Thickness (m)", "Insulation Density (kg/m³)"]
    data_sec = [
        ["sec_id_1", 0.1143, 0.00602, "pipe", 1000.0, 0.025, 200.0]
    ]
    
    for c_idx, h in enumerate(headers_sec, 1):
        ws_sec.cell(row=3, column=c_idx, value=h)
    for r_idx, row in enumerate(data_sec, 4):
        for c_idx, val in enumerate(row, 1):
            ws_sec.cell(row=r_idx, column=c_idx, value=val)
            if c_idx in (2, 3, 6):
                ws_sec.cell(row=r_idx, column=c_idx).number_format = '0.0000'
            elif c_idx in (5, 7):
                ws_sec.cell(row=r_idx, column=c_idx).number_format = '0.0'
                
    apply_table_styles(ws_sec, start_row=3, start_col=1, end_row=3+len(data_sec), end_col=len(headers_sec))
    autofit_columns(ws_sec)

    # 3. Nodes Sheet
    ws_nod = wb.create_sheet(title='Nodes')
    style_worksheet(ws_nod)
    ws_nod.cell(row=1, column=1, value="Node Coordinates").font = title_font
    
    headers_nod = ["Node ID", "X (m)", "Y (m)", "Z (m)"]
    data_nod = [
        ["0", 0.0, 0.0, 0.0],
        ["1", 3.0, 0.0, 0.0],
        ["2", 3.0, 3.0, 0.0],
        ["3", 5.0, 3.0, 0.0],
        ["4", 3.0, 3.0, 2.0]
    ]
    
    for c_idx, h in enumerate(headers_nod, 1):
        ws_nod.cell(row=3, column=c_idx, value=h)
    for r_idx, row in enumerate(data_nod, 4):
        for c_idx, val in enumerate(row, 1):
            ws_nod.cell(row=r_idx, column=c_idx, value=val)
            if c_idx > 1:
                ws_nod.cell(row=r_idx, column=c_idx).number_format = '0.000'
                
    apply_table_styles(ws_nod, start_row=3, start_col=1, end_row=3+len(data_nod), end_col=len(headers_nod))
    autofit_columns(ws_nod)

    # 4. Elements Sheet
    ws_elem = wb.create_sheet(title='Elements')
    style_worksheet(ws_elem)
    ws_elem.cell(row=1, column=1, value="Elements Connectivity").font = title_font
    
    headers_elem = ["Element ID", "Node A", "Node B", "Type", "Bend Radius (m)", "Material ID", "Section ID"]
    data_elem = [
        [0, "0", "1", "pipe", "", "mat_id_1", "sec_id_1"],
        [1, "1", "2", "bend", 0.17145, "mat_id_1", "sec_id_1"],
        [2, "2", "3", "pipe", "", "mat_id_1", "sec_id_1"],
        [3, "2", "4", "pipe", "", "mat_id_1", "sec_id_1"]
    ]
    
    for c_idx, h in enumerate(headers_elem, 1):
        ws_elem.cell(row=3, column=c_idx, value=h)
    for r_idx, row in enumerate(data_elem, 4):
        for c_idx, val in enumerate(row, 1):
            ws_elem.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 5 and val != "":
                ws_elem.cell(row=r_idx, column=c_idx).number_format = '0.0000'
                
    apply_table_styles(ws_elem, start_row=3, start_col=1, end_row=3+len(data_elem), end_col=len(headers_elem))
    autofit_columns(ws_elem)

    # 5. Boundary Conditions Sheet
    ws_bc = wb.create_sheet(title='Boundary_Conditions')
    style_worksheet(ws_bc)
    ws_bc.cell(row=1, column=1, value="Boundary Conditions").font = title_font
    ws_bc.cell(row=2, column=1, value="Enter TRUE for rigid fixity, FALSE/blank for free, or a number for spring stiffness (N/m).").font = Font(name='Segoe UI', size=9, italic=True)
    
    headers_bc = ["Node ID", "tx", "ty", "tz", "rx", "ry", "rz"]
    data_bc = [
        ["0", True, True, True, True, True, True],
        ["3", "", 50000.0, "", "", "", ""],
        ["4", True, "", True, "", "", ""]
    ]
    
    for c_idx, h in enumerate(headers_bc, 1):
        ws_bc.cell(row=4, column=c_idx, value=h)
    for r_idx, row in enumerate(data_bc, 5):
        for c_idx, val in enumerate(row, 1):
            ws_bc.cell(row=r_idx, column=c_idx, value=val)
            if c_idx > 1 and isinstance(val, (int, float)):
                ws_bc.cell(row=r_idx, column=c_idx).number_format = '#,##0.0'
                
    apply_table_styles(ws_bc, start_row=4, start_col=1, end_row=4+len(data_bc), end_col=len(headers_bc))
    autofit_columns(ws_bc)

    # 6. Loads Sheet
    ws_ld = wb.create_sheet(title='Loads')
    style_worksheet(ws_ld)
    ws_ld.cell(row=1, column=1, value="Load Parameters").font = title_font
    
    headers_ld = ["Parameter Name", "Value"]
    data_ld = [
        ["global_gravity_x", 0.0],
        ["global_gravity_y", -9.81],
        ["global_gravity_z", 0.0],
        ["global_internal_pressure", 2.0e6],
        ["global_temperature_change", 120.0],
        ["occasional_g_x", 0.0],
        ["occasional_g_y", 0.0],
        ["occasional_g_z", 0.0]
    ]
    
    for c_idx, h in enumerate(headers_ld, 1):
        ws_ld.cell(row=3, column=c_idx, value=h)
    for r_idx, row in enumerate(data_ld, 4):
        for c_idx, val in enumerate(row, 1):
            ws_ld.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 2:
                if "pressure" in row[0]:
                    ws_ld.cell(row=r_idx, column=c_idx).number_format = '#,##0.0'
                elif "gravity" in row[0] or "g_x" in row[0] or "g_y" in row[0] or "g_z" in row[0]:
                    ws_ld.cell(row=r_idx, column=c_idx).number_format = '0.00'
                else:
                    ws_ld.cell(row=r_idx, column=c_idx).number_format = '0.0'
                    
    apply_table_styles(ws_ld, start_row=3, start_col=1, end_row=3+len(data_ld), end_col=len(headers_ld))
    autofit_columns(ws_ld)

    # Save workbook
    wb.save(filepath)
    print(f"Excel template generated successfully at {filepath}")

def read_excel_inputs(filepath: str) -> dict:
    """Reads input data from an Excel workbook and parses it into the FEAEngine JSON schema format."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found at {filepath}")
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 1. Materials
    ws_mat = wb['Materials']
    materials = {}
    for row in range(4, ws_mat.max_row + 1):
        mat_id = ws_mat.cell(row=row, column=1).value
        if mat_id is not None and str(mat_id).strip() != "":
            materials[str(mat_id)] = {
                "E": float(ws_mat.cell(row=row, column=2).value),
                "G": float(ws_mat.cell(row=row, column=3).value),
                "alpha": float(ws_mat.cell(row=row, column=4).value),
                "yield_strength": float(ws_mat.cell(row=row, column=5).value),
                "Sc": float(ws_mat.cell(row=row, column=6).value),
                "Sh": float(ws_mat.cell(row=row, column=7).value),
            }
            
    # 2. Sections
    ws_sec = wb['Sections']
    sections = {}
    for row in range(4, ws_sec.max_row + 1):
        sec_id = ws_sec.cell(row=row, column=1).value
        if sec_id is not None and str(sec_id).strip() != "":
            sections[str(sec_id)] = {
                "OD": float(ws_sec.cell(row=row, column=2).value),
                "wall_thickness": float(ws_sec.cell(row=row, column=3).value),
                "type": str(ws_sec.cell(row=row, column=4).value).strip().lower(),
                "fluid_density": float(ws_sec.cell(row=row, column=5).value or 0.0),
                "insulation_thickness": float(ws_sec.cell(row=row, column=6).value or 0.0),
                "insulation_density": float(ws_sec.cell(row=row, column=7).value or 0.0),
            }
            
    # 3. Nodes
    ws_nod = wb['Nodes']
    nodes = {}
    for row in range(4, ws_nod.max_row + 1):
        nod_id = ws_nod.cell(row=row, column=1).value
        if nod_id is not None and str(nod_id).strip() != "":
            nodes[str(nod_id)] = [
                float(ws_nod.cell(row=row, column=2).value),
                float(ws_nod.cell(row=row, column=3).value),
                float(ws_nod.cell(row=row, column=4).value),
            ]
            
    # 4. Elements
    ws_elem = wb['Elements']
    elements = []
    for row in range(4, ws_elem.max_row + 1):
        elem_id_val = ws_elem.cell(row=row, column=1).value
        if elem_id_val is not None and str(elem_id_val).strip() != "":
            elem_data = {
                "id": int(elem_id_val),
                "node_A": str(ws_elem.cell(row=row, column=2).value),
                "node_B": str(ws_elem.cell(row=row, column=3).value),
                "type": str(ws_elem.cell(row=row, column=4).value).strip().lower(),
                "material": str(ws_elem.cell(row=row, column=6).value),
                "section": str(ws_elem.cell(row=row, column=7).value),
            }
            
            bend_rad_val = ws_elem.cell(row=row, column=5).value
            if bend_rad_val is not None and str(bend_rad_val).strip() != "":
                elem_data["bend_radius"] = float(bend_rad_val)
                
            elements.append(elem_data)
            
    # 5. Boundary Conditions
    ws_bc = wb['Boundary_Conditions']
    boundary_conditions = {}
    for row in range(5, ws_bc.max_row + 1):
        nod_id = ws_bc.cell(row=row, column=1).value
        if nod_id is not None and str(nod_id).strip() != "":
            bc_entry = {}
            for col_idx, dof in enumerate(["tx", "ty", "tz", "rx", "ry", "rz"], 2):
                val = ws_bc.cell(row=row, column=col_idx).value
                parsed = parse_bc_value(val)
                if parsed is not False:
                    bc_entry[dof] = parsed
            if bc_entry:
                boundary_conditions[str(nod_id)] = bc_entry
                
    # 6. Loads
    ws_ld = wb['Loads']
    raw_loads = {}
    for row in range(4, ws_ld.max_row + 1):
        param = ws_ld.cell(row=row, column=1).value
        if param is not None and str(param).strip() != "":
            raw_loads[str(param).strip()] = ws_ld.cell(row=row, column=2).value
            
    loads = {
        "global_gravity": [
            float(raw_loads.get("global_gravity_x", 0.0)),
            float(raw_loads.get("global_gravity_y", -9.81)),
            float(raw_loads.get("global_gravity_z", 0.0))
        ],
        "global_internal_pressure": float(raw_loads.get("global_internal_pressure", 0.0)),
        "global_temperature_change": float(raw_loads.get("global_temperature_change", 0.0)),
        "occasional_g": [
            float(raw_loads.get("occasional_g_x", 0.0)),
            float(raw_loads.get("occasional_g_y", 0.0)),
            float(raw_loads.get("occasional_g_z", 0.0))
        ]
    }
    
    return {
        "materials": materials,
        "sections": sections,
        "nodes": nodes,
        "elements": elements,
        "boundary_conditions": boundary_conditions,
        "loads": loads
    }

def write_excel_results(filepath: str, summary: dict):
    """Writes numerical solver results to new formatted tabs."""
    wb = openpyxl.load_workbook(filepath)
    
    # Overwrite if exists
    for s_name in ["Displacements_Report", "Stress_Compliance_Report"]:
        if s_name in wb.sheetnames:
            wb.remove(wb[s_name])
            
    title_font = Font(name='Segoe UI', size=14, bold=True, color='1F4E79')
    
    # 1. Displacements
    ws_disp = wb.create_sheet(title="Displacements_Report")
    style_worksheet(ws_disp)
    ws_disp.cell(row=1, column=1, value="Nodal Displacements Report").font = title_font
    
    headers_disp = ["Node ID", "Load Case", "Dx (m)", "Dy (m)", "Dz (m)", "Rx (rad)", "Ry (rad)", "Rz (rad)"]
    for c_idx, h in enumerate(headers_disp, 1):
        ws_disp.cell(row=3, column=c_idx, value=h)
        
    disp_rows = []
    sorted_node_ids = sorted(summary['nodes'].keys(), key=lambda x: int(x) if x.isdigit() else x)
    for nid in sorted_node_ids:
        cases = summary['nodes'][nid]
        for case in ["Weight", "Thermal", "Operating"]:
            if case in cases:
                row_data = [nid, case] + [float(val) for val in cases[case]]
                disp_rows.append(row_data)
                
    for r_idx, row in enumerate(disp_rows, 4):
        for c_idx, val in enumerate(row, 1):
            ws_disp.cell(row=r_idx, column=c_idx, value=val)
            if c_idx > 2:
                ws_disp.cell(row=r_idx, column=c_idx).number_format = '0.000000'
                
    apply_table_styles(ws_disp, start_row=3, start_col=1, end_row=3+len(disp_rows), end_col=len(headers_disp))
    autofit_columns(ws_disp)
    
    # 2. Stress Compliance
    ws_stress = wb.create_sheet(title="Stress_Compliance_Report")
    style_worksheet(ws_stress)
    ws_stress.cell(row=1, column=1, value="Element Stress & Compliance Report").font = title_font
    
    headers_stress = [
        "Element ID", "Type", "Flexibility Factor", "SIF In-Plane", "SIF Out-of-Plane",
        "SUS Stress (Pa)", "SUS Allow (Pa)", "SUS Ratio", 
        "EXP Stress (Pa)", "EXP Allow (Pa)", "EXP Ratio", 
        "OCC Stress (Pa)", "OCC Allow (Pa)", "OCC Ratio", 
        "Max Stress Ratio", "Status"
    ]
    for c_idx, h in enumerate(headers_stress, 1):
        ws_stress.cell(row=3, column=c_idx, value=h)
        
    stress_rows = []
    sorted_elem_ids = sorted(summary['elements'].keys(), key=lambda x: int(x) if isinstance(x, int) or str(x).isdigit() else x)
    for eid in sorted_elem_ids:
        el_res = summary['elements'][eid]
        status_str = "PASS" if el_res['compliance_pass'] else "FAIL"
        row_data = [
            eid,
            el_res['type'],
            el_res['k_factor'],
            el_res['SIF_in'],
            el_res['SIF_out'],
            el_res['sustained_stress'],
            el_res['sustained_allowable'],
            el_res['sustained_ratio'],
            el_res['expansion_stress'],
            el_res['expansion_allowable'],
            el_res['expansion_ratio'],
            el_res['occasional_stress'],
            el_res['occasional_allowable'],
            el_res['occasional_ratio'],
            el_res['max_stress_ratio'],
            status_str
        ]
        stress_rows.append(row_data)
        
    for r_idx, row in enumerate(stress_rows, 4):
        for c_idx, val in enumerate(row, 1):
            cell = ws_stress.cell(row=r_idx, column=c_idx, value=val)
            if c_idx in (3, 4, 5):
                cell.number_format = '0.00'
            elif c_idx in (6, 7, 9, 10, 12, 13):
                cell.number_format = '0.00E+00'
            elif c_idx in (8, 11, 14, 15):
                cell.number_format = '0.00'
            elif c_idx == 16:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if val == "PASS":
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='006100')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=10, bold=True, color='9C0006')
                    
    apply_table_styles(ws_stress, start_row=3, start_col=1, end_row=3+len(stress_rows), end_col=len(headers_stress))
    
    # Conditional formatting on column 15 (Max Stress Ratio)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(name='Segoe UI', size=10, color='9C0006', bold=True)
    ws_stress.conditional_formatting.add(
        f'O4:O{3+len(stress_rows)}',
        CellIsRule(operator='greaterThan', formula=['1.0'], stopIfTrue=True, fill=red_fill, font=red_font)
    )
    
    autofit_columns(ws_stress)
    wb.save(filepath)
    print(f"Results successfully saved and formatted in Excel workbook: {filepath}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Spreadsheet-based solver interface for Pipe Stress Analyzer.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--generate", metavar="TEMPLATE_NAME", help="Path to generate a clean Excel template file.")
    group.add_argument("--solve", metavar="INPUT_EXCEL", help="Path to the Excel file containing user inputs to analyze.")
    
    parser.add_argument("--output", metavar="OUTPUT_EXCEL", help="Output Excel filename to save results into (defaults to updating the input Excel file).")
    
    args = parser.parse_args()
    
    if args.generate:
        generate_template(args.generate)
    elif args.solve:
        print(f"Reading Excel inputs from: {args.solve} ...")
        try:
            input_data = read_excel_inputs(args.solve)
        except Exception as e:
            print(f"Error parsing Excel spreadsheet: {e}")
            sys.exit(1)
            
        print("Initializing FEA Structural Stress Solver...")
        engine = FEAEngine(input_data)
        
        print("Running weight, pressure, thermal expansion, and occasional compliance loops...")
        engine.solve()
        
        summary = engine.get_summary()
        
        output_file = args.output if args.output else args.solve
        print(f"Saving solver results back to: {output_file} ...")
        try:
            write_excel_results(output_file, summary)
        except Exception as e:
            print(f"Error writing solver results back to Excel: {e}")
            sys.exit(1)
            
        print("Stress compliance complete! Summary:")
        print(f"  Compliance Pass: {not summary['compliance_warning']}")
        print(f"  Max Stress Ratio: {max(el['max_stress_ratio'] for el in summary['elements'].values()):.2f}")
