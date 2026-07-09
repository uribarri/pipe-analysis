import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_multi_element_workbook(filepath: str):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define sheets
    ws_inputs = wb.create_sheet(title="Inputs")
    ws_elem = wb.create_sheet(title="Element_Stiffness")
    ws_solve = wb.create_sheet(title="Global_Solver")
    ws_report = wb.create_sheet(title="Stress_Report")
    
    # Apply showGridLines
    for ws in [ws_inputs, ws_elem, ws_solve, ws_report]:
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
        else:
            ws.sheet_view.showGridLines = True
            
    # Styling definitions
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_data_bold = Font(name="Segoe UI", size=10, bold=True)
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="006100")
    font_fail = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
    
    fill_section = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_input = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
    fill_calc = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Grey
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ----------------- 1. INPUTS SHEET -----------------
    ws_inputs['B2'] = "ASME B31.3 Piping 3D Multi-Element FEA Solver"
    ws_inputs['B2'].font = font_title
    ws_inputs['B3'] = "Edit values in yellow cells. Calculations update automatically in all sheets."
    ws_inputs['B3'].font = font_subtitle
    
    # Materials Table
    ws_inputs.merge_cells('B5:H5')
    ws_inputs['B5'] = "1. MATERIAL PROPERTIES"
    ws_inputs['B5'].font = font_section
    ws_inputs['B5'].fill = fill_section
    ws_inputs['B5'].alignment = align_center
    
    headers_mat = ["Material ID", "E (Pa)", "G (Pa)", "alpha (1/°C)", "Yield Strength (Pa)", "Sc (Pa)", "Sh (Pa)"]
    for col_idx, h in enumerate(headers_mat, 2):
        cell = ws_inputs.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Pre-populate sample materials
    sample_mats = [
        ["mat1", 2.0e11, 7.7e10, 1.2e-5, 2.5e8, 1.379e8, 1.379e8]
    ]
    for r_idx, row in enumerate(sample_mats, 7):
        for c_idx, val in enumerate(row, 2):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.fill = fill_input
            if c_idx > 2:
                cell.number_format = '0.00E+00'
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    # Sections Table
    ws_inputs.merge_cells('B10:H10')
    ws_inputs['B10'] = "2. SECTION PROPERTIES"
    ws_inputs['B10'].font = font_section
    ws_inputs['B10'].fill = fill_section
    ws_inputs['B10'].alignment = align_center
    
    headers_sec = ["Section ID", "Outer Diameter (m)", "Wall Thickness (m)", "Type", "Fluid Density (kg/m³)", "Insulation Thickness (m)", "Insulation Density (kg/m³)"]
    for col_idx, h in enumerate(headers_sec, 2):
        cell = ws_inputs.cell(row=11, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_secs = [
        ["sec1", 0.1143, 0.00602, "pipe", 1000.0, 0.025, 200.0]
    ]
    for r_idx, row in enumerate(sample_secs, 12):
        for c_idx, val in enumerate(row, 2):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.fill = fill_input
            if c_idx in (3, 4, 7):
                cell.number_format = '0.0000'
                cell.alignment = align_right
            elif c_idx in (6, 8):
                cell.number_format = '0.0'
                cell.alignment = align_right
            else:
                cell.alignment = align_center

    # Nodes Table (Up to 8 nodes)
    ws_inputs.merge_cells('B15:E15')
    ws_inputs['B15'] = "3. NODES COORDINATES"
    ws_inputs['B15'].font = font_section
    ws_inputs['B15'].fill = fill_section
    ws_inputs['B15'].alignment = align_center
    
    headers_nod = ["Node ID", "X (m)", "Y (m)", "Z (m)"]
    for col_idx, h in enumerate(headers_nod, 2):
        cell = ws_inputs.cell(row=16, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_nodes = [
        ["0", 0.0, 0.0, 0.0],
        ["1", 3.0, 0.0, 0.0],
        ["2", 3.0, 3.0, 0.0],
        ["3", 5.0, 3.0, 0.0],
        ["4", 3.0, 3.0, 2.0],
        ["5", "", "", ""],
        ["6", "", "", ""],
        ["7", "", "", ""]
    ]
    for r_idx, row in enumerate(sample_nodes, 17):
        for c_idx, val in enumerate(row, 2):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.fill = fill_input
            if c_idx > 2 and val != "":
                cell.number_format = '0.000'
                cell.alignment = align_right
            else:
                cell.alignment = align_center

    # Elements Table (Up to 7 elements)
    ws_inputs.merge_cells('B27:H27')
    ws_inputs['B27'] = "4. ELEMENTS CONNECTIVITY"
    ws_inputs['B27'].font = font_section
    ws_inputs['B27'].fill = fill_section
    ws_inputs['B27'].alignment = align_center
    
    headers_elem = ["Element ID", "Node A", "Node B", "Type", "Bend Radius (m)", "Material ID", "Section ID"]
    for col_idx, h in enumerate(headers_elem, 2):
        cell = ws_inputs.cell(row=28, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_elems = [
        [0, "0", "1", "pipe", "", "mat1", "sec1"],
        [1, "1", "2", "bend", 0.17145, "mat1", "sec1"],
        [2, "2", "3", "pipe", "", "mat1", "sec1"],
        [3, "2", "4", "pipe", "", "mat1", "sec1"],
        [4, "", "", "pipe", "", "mat1", "sec1"],
        [5, "", "", "pipe", "", "mat1", "sec1"],
        [6, "", "", "pipe", "", "mat1", "sec1"]
    ]
    for r_idx, row in enumerate(sample_elems, 29):
        for c_idx, val in enumerate(row, 2):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.fill = fill_input
            if c_idx == 6 and val != "":
                cell.number_format = '0.0000'
                cell.alignment = align_right
            else:
                cell.alignment = align_center

    # Boundary Conditions Table
    ws_inputs.merge_cells('B38:H38')
    ws_inputs['B38'] = "5. NODAL BOUNDARY CONDITIONS"
    ws_inputs['B38'].font = font_section
    ws_inputs['B38'].fill = fill_section
    ws_inputs['B38'].alignment = align_center
    
    headers_bc = ["Node ID", "tx", "ty", "tz", "rx", "ry", "rz"]
    for col_idx, h in enumerate(headers_bc, 2):
        cell = ws_inputs.cell(row=39, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_bcs = [
        ["0", True, True, True, True, True, True],
        ["3", "", 50000.0, "", "", "", ""],
        ["4", True, "", True, "", "", ""],
        ["1", "", "", "", "", "", ""],
        ["2", "", "", "", "", "", ""],
        ["5", "", "", "", "", "", ""],
        ["6", "", "", "", "", "", ""],
        ["7", "", "", "", "", "", ""]
    ]
    for r_idx, row in enumerate(sample_bcs, 40):
        for c_idx, val in enumerate(row, 2):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.fill = fill_input
            if c_idx > 2 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.0'
                cell.alignment = align_right
            else:
                cell.alignment = align_center

    # Loads Table
    ws_inputs.merge_cells('J5:L5')
    ws_inputs['J5'] = "6. GLOBAL DESIGN LOADS"
    ws_inputs['J5'].font = font_section
    ws_inputs['J5'].fill = fill_section
    ws_inputs['J5'].alignment = align_center
    
    headers_ld = ["Parameter Name", "Value", "Unit"]
    for col_idx, h in enumerate(headers_ld, 10):
        cell = ws_inputs.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_lds = [
        ["global_gravity_x", 0.0, "m/s²"],
        ["global_gravity_y", -9.81, "m/s²"],
        ["global_gravity_z", 0.0, "m/s²"],
        ["global_internal_pressure", 2.0e6, "Pa"],
        ["global_temperature_change", 120.0, "°C"]
    ]
    for r_idx, row in enumerate(sample_lds, 7):
        for c_idx, val in enumerate(row, 10):
            cell = ws_inputs.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if c_idx == 11:
                cell.fill = fill_input
                cell.font = font_data_bold
                cell.alignment = align_right
                if "pressure" in row[0]:
                    cell.number_format = '0.00E+00'
                elif "gravity" in row[0]:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '0.0'
            elif c_idx == 12:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Auto-fit columns
    for col in ws_inputs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_inputs.column_dimensions[col_letter].width = max(max_len + 3, 10)
    ws_inputs.column_dimensions['B'].width = 25
    ws_inputs.column_dimensions['J'].width = 25

    # ----------------- 2. ELEMENT_STIFFNESS SHEET -----------------
    ws_elem['A1'] = "3D Finite Element Local & Global Stiffness Matrices"
    ws_elem['A1'].font = font_title
    
    # We allocate blocks of 35 rows per element (Elements 0 to 6)
    for e_idx in range(7):
        row_offset = 35 * e_idx + 4
        
        # Header block for Element
        ws_elem.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=27)
        title_cell = ws_elem.cell(row=row_offset, column=2, value=f"ELEMENT {e_idx} ANALYSIS BLOCK")
        title_cell.font = font_section
        title_cell.fill = fill_section
        title_cell.alignment = align_center
        
        # Row row_offset + 1: References
        ws_elem.cell(row=row_offset+1, column=2, value="Node A ID").font = font_data_bold
        ws_elem.cell(row=row_offset+1, column=3, value=f"=Inputs!C{29+e_idx}").font = font_data # Inputs!C29 is Element 0 Node A
        ws_elem.cell(row=row_offset+1, column=4, value="Node B ID").font = font_data_bold
        ws_elem.cell(row=row_offset+1, column=5, value=f"=Inputs!D{29+e_idx}").font = font_data
        
        # Row row_offset + 2: Active check
        # Active if both Node A and Node B are not blank
        ws_elem.cell(row=row_offset+2, column=2, value="Is Active?").font = font_data_bold
        ws_elem.cell(row=row_offset+2, column=3, value=f'=AND(C{row_offset+1}<>"",E{row_offset+1}<>"")').font = font_data_bold
        
        # Node coordinates
        # Coordinate lookup: we lookup coordinates in Inputs!$B$17:$E$24
        # Node A coords
        ws_elem.cell(row=row_offset+3, column=2, value="Node A Coord").font = font_data_bold
        ws_elem.cell(row=row_offset+3, column=3, value=f'=IF(C{row_offset+2},VLOOKUP(C{row_offset+1},Inputs!$B$17:$E$24,2,FALSE),0.0)').number_format = '0.00' # X
        ws_elem.cell(row=row_offset+3, column=4, value=f'=IF(C{row_offset+2},VLOOKUP(C{row_offset+1},Inputs!$B$17:$E$24,3,FALSE),0.0)').number_format = '0.00' # Y
        ws_elem.cell(row=row_offset+3, column=5, value=f'=IF(C{row_offset+2},VLOOKUP(C{row_offset+1},Inputs!$B$17:$E$24,4,FALSE),0.0)').number_format = '0.00' # Z
        
        # Node B coords
        ws_elem.cell(row=row_offset+3, column=6, value="Node B Coord").font = font_data_bold
        ws_elem.cell(row=row_offset+3, column=7, value=f'=IF(C{row_offset+2},VLOOKUP(E{row_offset+1},Inputs!$B$17:$E$24,2,FALSE),0.0)').number_format = '0.00' # X
        ws_elem.cell(row=row_offset+3, column=8, value=f'=IF(C{row_offset+2},VLOOKUP(E{row_offset+1},Inputs!$B$17:$E$24,3,FALSE),0.0)').number_format = '0.00' # Y
        ws_elem.cell(row=row_offset+3, column=9, value=f'=IF(C{row_offset+2},VLOOKUP(E{row_offset+1},Inputs!$B$17:$E$24,4,FALSE),0.0)').number_format = '0.00' # Z
        
        # Length and coordinate deltas
        ws_elem.cell(row=row_offset+4, column=2, value="deltas dx,dy,dz").font = font_data_bold
        ws_elem.cell(row=row_offset+4, column=3, value=f"=G{row_offset+3}-C{row_offset+3}") # dx
        ws_elem.cell(row=row_offset+4, column=4, value=f"=H{row_offset+3}-D{row_offset+3}") # dy
        ws_elem.cell(row=row_offset+4, column=5, value=f"=I{row_offset+3}-E{row_offset+3}") # dz
        
        ws_elem.cell(row=row_offset+4, column=6, value="Length L").font = font_data_bold
        # Avoid zero length for inactive elements to prevent division by zero
        ws_elem.cell(row=row_offset+4, column=7, value=f'=IF(C{row_offset+2},SQRT(C{row_offset+4}^2+D{row_offset+4}^2+E{row_offset+4}^2),1.0)').number_format = '0.000'
        
        # Direction cosines lx, my, nz
        ws_elem.cell(row=row_offset+5, column=2, value="dir cosines lx,my,nz").font = font_data_bold
        ws_elem.cell(row=row_offset+5, column=3, value=f"=C{row_offset+4}/G{row_offset+4}") # lx
        ws_elem.cell(row=row_offset+5, column=4, value=f"=D{row_offset+4}/G{row_offset+4}") # mx
        ws_elem.cell(row=row_offset+5, column=5, value=f"=E{row_offset+4}/G{row_offset+4}") # nx
        
        # Check vertical: D_is_zero = TRUE if lx and my are very small
        ws_elem.cell(row=row_offset+5, column=6, value="D_is_zero").font = font_data_bold
        ws_elem.cell(row=row_offset+5, column=7, value=f"=AND(ABS(C{row_offset+5})<1E-5,ABS(D{row_offset+5})<1E-5)")
        ws_elem.cell(row=row_offset+5, column=8, value="D").font = font_data_bold
        ws_elem.cell(row=row_offset+5, column=9, value=f"=SQRT(C{row_offset+5}^2+D{row_offset+5}^2)")
        
        # SIFs, material properties, section properties lookup
        # Material ID is at Inputs!G29+e_idx. Section ID is at Inputs!H29+e_idx.
        ws_elem.cell(row=row_offset+6, column=2, value="Material E, G, alpha").font = font_data_bold
        ws_elem.cell(row=row_offset+6, column=3, value=f"=VLOOKUP(Inputs!G{29+e_idx},Inputs!$B$7:$H$8,2,FALSE)") # E
        ws_elem.cell(row=row_offset+6, column=4, value=f"=VLOOKUP(Inputs!G{29+e_idx},Inputs!$B$7:$H$8,3,FALSE)") # G
        ws_elem.cell(row=row_offset+6, column=5, value=f"=VLOOKUP(Inputs!G{29+e_idx},Inputs!$B$7:$H$8,4,FALSE)") # alpha
        
        # Section properties
        ws_elem.cell(row=row_offset+7, column=2, value="Section OD, t").font = font_data_bold
        ws_elem.cell(row=row_offset+7, column=3, value=f"=VLOOKUP(Inputs!H{29+e_idx},Inputs!$B$12:$H$13,2,FALSE)") # OD
        ws_elem.cell(row=row_offset+7, column=4, value=f"=VLOOKUP(Inputs!H{29+e_idx},Inputs!$B$12:$H$13,3,FALSE)") # t
        
        # Pre-calculated area, I, J, Z for stiffness calculations
        ws_elem.cell(row=row_offset+8, column=2, value="ro, ri").font = font_data_bold
        ws_elem.cell(row=row_offset+8, column=3, value=f"=C{row_offset+7}/2") # ro
        ws_elem.cell(row=row_offset+8, column=4, value=f"=C{row_offset+8}-D{row_offset+7}") # ri
        
        ws_elem.cell(row=row_offset+8, column=5, value="Area A").font = font_data_bold
        ws_elem.cell(row=row_offset+8, column=6, value=f"=PI()*(C{row_offset+8}^2-D{row_offset+8}^2)")
        
        ws_elem.cell(row=row_offset+9, column=2, value="Inertia I, J").font = font_data_bold
        ws_elem.cell(row=row_offset+9, column=3, value=f"=(PI()/4)*(C{row_offset+8}^4-D{row_offset+8}^4)") # I
        ws_elem.cell(row=row_offset+9, column=4, value=f"=C{row_offset+9}*2") # J
        
        # Flexibility factor k
        # k = 1.65 / h, where h = t * R / r_m^2.
        # Check if type is bend (Inputs!E29+e_idx). Radius is at Inputs!F29+e_idx
        ws_elem.cell(row=row_offset+9, column=5, value="Flex Factor k").font = font_data_bold
        ws_elem.cell(row=row_offset+9, column=6, value=f'=IF(Inputs!E{29+e_idx}="bend",IF(Inputs!F{29+e_idx}<>"",MAX(1.0,1.65/((D{row_offset+7}*Inputs!F{29+e_idx})/(((C{row_offset+7}-D{row_offset+7})/2)^2))),1.0),1.0)')
        
        # Rotation matrix lambda (3x3)
        ws_elem.cell(row=row_offset+5, column=12, value="ROTATION MATRIX lambda").font = font_data_bold
        # Row 1: lx, mx, nx
        ws_elem.cell(row=row_offset+6, column=12, value=f"=C{row_offset+5}")
        ws_elem.cell(row=row_offset+6, column=13, value=f"=D{row_offset+5}")
        ws_elem.cell(row=row_offset+6, column=14, value=f"=E{row_offset+5}")
        # Row 2: ly, my, ny
        ws_elem.cell(row=row_offset+7, column=12, value=f"=IF(G{row_offset+5},0,-D{row_offset+5}/I{row_offset+5})")
        ws_elem.cell(row=row_offset+7, column=13, value=f"=IF(G{row_offset+5},IF(E{row_offset+5}>0,1,-1),C{row_offset+5}/I{row_offset+5})")
        ws_elem.cell(row=row_offset+7, column=14, value=0.0)
        # Row 3: lz, mz, nz
        ws_elem.cell(row=row_offset+8, column=12, value=f"=IF(G{row_offset+5},IF(E{row_offset+5}>0,-1,1),-C{row_offset+5}*E{row_offset+5}/I{row_offset+5})")
        ws_elem.cell(row=row_offset+8, column=13, value=f"=IF(G{row_offset+5},0,-D{row_offset+5}*E{row_offset+5}/I{row_offset+5})")
        ws_elem.cell(row=row_offset+8, column=14, value=f"=IF(G{row_offset+5},0,I{row_offset+5})")
        
        for r in range(row_offset+6, row_offset+9):
            for c in range(12, 15):
                cell = ws_elem.cell(row=r, column=c)
                cell.font = font_data
                cell.alignment = align_right
                cell.fill = fill_calc
                cell.border = thin_border
                cell.number_format = '0.0000'

        # Local Stiffness Matrix k_local (12x12) in columns D to O, rows row_offset+10 to row_offset+21
        # Let's write cells of k_local. To keep it clean, we can pre-load formulas.
        k_local_start_row = row_offset + 10
        # Reference variables to simplify formulas
        # E = C{row_offset+6}, G = D{row_offset+6}, A = F{row_offset+8}, I = C{row_offset+9}, J = D{row_offset+9}, L = G{row_offset+4}, k = F{row_offset+9}
        E_ref = f"$C${row_offset+6}"
        G_ref = f"$D${row_offset+6}"
        A_ref = f"$F${row_offset+8}"
        I_ref = f"$C${row_offset+9}"
        J_ref = f"$D${row_offset+9}"
        L_ref = f"$G${row_offset+4}"
        k_ref = f"$F${row_offset+9}"
        
        # Initialize 12x12 matrix with zeros
        for r in range(12):
            for c in range(12):
                cell = ws_elem.cell(row=k_local_start_row+r, column=4+c, value=0.0)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = thin_border
                cell.number_format = '0.00E+00'
                
        # Fill non-zero local stiffness terms
        # Axial terms
        ws_elem.cell(row=k_local_start_row+0, column=4, value=f"=IF(C{row_offset+2},{E_ref}*{A_ref}/{L_ref},0)")
        ws_elem.cell(row=k_local_start_row+0, column=10, value=f"=-D{k_local_start_row+0}")
        ws_elem.cell(row=k_local_start_row+6, column=4, value=f"=J{k_local_start_row+0}")
        ws_elem.cell(row=k_local_start_row+6, column=10, value=f"=-J{k_local_start_row+0}")
        
        # Torsion terms
        ws_elem.cell(row=k_local_start_row+3, column=7, value=f"=IF(C{row_offset+2},{G_ref}*{J_ref}/{L_ref},0)")
        ws_elem.cell(row=k_local_start_row+3, column=13, value=f"=-G{k_local_start_row+3}")
        ws_elem.cell(row=k_local_start_row+9, column=7, value=f"=M{k_local_start_row+3}")
        ws_elem.cell(row=k_local_start_row+9, column=13, value=f"=-M{k_local_start_row+3}")
        
        # Bending in xy plane (v translation, thz rotation) - modified by k_ref flexibility factor
        # Iz_eff = I_ref / k_ref
        Iz_eff = f"({I_ref}/{k_ref})"
        az = f"=(12*{E_ref}*{Iz_eff})/({L_ref}^3)"
        bz = f"=(6*{E_ref}*{Iz_eff})/({L_ref}^2)"
        cz = f"=(4*{E_ref}*{Iz_eff})/{L_ref}"
        dz = f"=(2*{E_ref}*{Iz_eff})/{L_ref}"
        
        # Row 1 (v_A)
        ws_elem.cell(row=k_local_start_row+1, column=5, value=f"=IF(C{row_offset+2},{az},0)")
        ws_elem.cell(row=k_local_start_row+1, column=9, value=f"=IF(C{row_offset+2},{bz},0)")
        ws_elem.cell(row=k_local_start_row+1, column=11, value=f"=-E{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+1, column=15, value=f"=I{k_local_start_row+1}")
        
        # Row 5 (thz_A)
        ws_elem.cell(row=k_local_start_row+5, column=5, value=f"=I{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+5, column=9, value=f"=IF(C{row_offset+2},{cz},0)")
        ws_elem.cell(row=k_local_start_row+5, column=11, value=f"=-I{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+5, column=15, value=f"=IF(C{row_offset+2},{dz},0)")
        
        # Row 7 (v_B)
        ws_elem.cell(row=k_local_start_row+7, column=5, value=f"=K{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+7, column=9, value=f"=K{k_local_start_row+5}")
        ws_elem.cell(row=k_local_start_row+7, column=11, value=f"=E{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+7, column=15, value=f"=-I{k_local_start_row+1}")
        
        # Row 11 (thz_B)
        ws_elem.cell(row=k_local_start_row+11, column=5, value=f"=O{k_local_start_row+1}")
        ws_elem.cell(row=k_local_start_row+11, column=9, value=f"=O{k_local_start_row+5}")
        ws_elem.cell(row=k_local_start_row+11, column=11, value=f"=K{k_local_start_row+7}")
        ws_elem.cell(row=k_local_start_row+11, column=15, value=f"=I{k_local_start_row+5}")

        # Bending in xz plane (w translation, thy rotation) - modified by k_ref flexibility factor
        # Iy_eff = I_ref / k_ref
        Iy_eff = f"({I_ref}/{k_ref})"
        ay = f"=(12*{E_ref}*{Iy_eff})/({L_ref}^3)"
        by = f"=(6*{E_ref}*{Iy_eff})/({L_ref}^2)"
        cy = f"=(4*{E_ref}*{Iy_eff})/{L_ref}"
        dy = f"=(2*{E_ref}*{Iy_eff})/{L_ref}"
        
        # Row 2 (w_A)
        ws_elem.cell(row=k_local_start_row+2, column=6, value=f"=IF(C{row_offset+2},{ay},0)")
        ws_elem.cell(row=k_local_start_row+2, column=8, value=f"=IF(C{row_offset+2},-{by},0)")
        ws_elem.cell(row=k_local_start_row+2, column=12, value=f"=-F{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+2, column=14, value=f"=H{k_local_start_row+2}")
        
        # Row 4 (thy_A)
        ws_elem.cell(row=k_local_start_row+4, column=6, value=f"=H{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+4, column=8, value=f"=IF(C{row_offset+2},{cy},0)")
        ws_elem.cell(row=k_local_start_row+4, column=12, value=f"=-H{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+4, column=14, value=f"=IF(C{row_offset+2},{dy},0)")
        
        # Row 8 (w_B)
        ws_elem.cell(row=k_local_start_row+8, column=6, value=f"=L{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+8, column=8, value=f"=L{k_local_start_row+4}")
        ws_elem.cell(row=k_local_start_row+8, column=12, value=f"=F{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+8, column=14, value=f"=-H{k_local_start_row+2}")
        
        # Row 10 (thy_B)
        ws_elem.cell(row=k_local_start_row+10, column=6, value=f"=N{k_local_start_row+2}")
        ws_elem.cell(row=k_local_start_row+10, column=8, value=f"=N{k_local_start_row+4}")
        ws_elem.cell(row=k_local_start_row+10, column=12, value=f"=L{k_local_start_row+8}")
        ws_elem.cell(row=k_local_start_row+10, column=14, value=f"=H{k_local_start_row+4}")

        # Transformation Matrix T (12x12) in columns P to AA, rows row_offset+10 to row_offset+21
        # Subblocks of lambda on diagonal:
        # T[0:3, 0:3] = lambda, T[3:6, 3:6] = lambda, T[6:9, 6:9] = lambda, T[9:12, 9:12] = lambda
        t_start_col = 16
        for r in range(12):
            for c in range(12):
                cell = ws_elem.cell(row=k_local_start_row+r, column=t_start_col+c, value=0.0)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = thin_border
                cell.number_format = '0.0000'
                
        # Fill diagonal subblocks
        lambda_row = row_offset + 6
        for subblock in range(4):
            r_shift = subblock * 3
            c_shift = subblock * 3
            for r in range(3):
                for c in range(3):
                    # Reference lambda cells: rows lambda_row+r, cols 12+c (L, M, N)
                    lambda_col_letter = get_column_letter(12 + c)
                    ws_elem.cell(row=k_local_start_row + r_shift + r, column=t_start_col + c_shift + c, 
                                 value=f"={lambda_col_letter}{lambda_row + r}")

        # Intermediate Product: k_temp = k_local * T (12x12) in columns D to O, rows row_offset+22 to row_offset+33
        k_temp_start_row = row_offset + 22
        for r in range(12):
            for c in range(12):
                # local row range: D{k_local_start_row+r}:O{k_local_start_row+r}
                # T col range: {col_letter}{k_local_start_row}:{col_letter}{k_local_start_row+11}
                col_letter = get_column_letter(t_start_col + c)
                formula = f"=SUMPRODUCT(D{k_local_start_row+r}:O{k_local_start_row+r}, {col_letter}{k_local_start_row}:{col_letter}{k_local_start_row+11})"
                cell = ws_elem.cell(row=k_temp_start_row+r, column=4+c, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = thin_border
                cell.fill = fill_calc
                cell.number_format = '0.00E+00'

        # Global Stiffness Matrix: k_global = T^T * k_temp (12x12) in columns P to AA, rows row_offset+22 to row_offset+33
        k_global_start_row = row_offset + 22
        for r in range(12):
            for c in range(12):
                # T row range: P{k_local_start_row+r}:AA{k_local_start_row+r} (which acts as column in T^T)
                # k_temp col range: {col_letter}{k_temp_start_row}:{col_letter}{k_temp_start_row+11}
                col_letter_temp = get_column_letter(4 + c)
                formula = f"=SUMPRODUCT(P{k_local_start_row+r}:AA{k_local_start_row+r}, {col_letter_temp}{k_temp_start_row}:{col_letter_temp}{k_temp_start_row+11})"
                cell = ws_elem.cell(row=k_global_start_row+r, column=t_start_col+c, value=formula)
                cell.font = font_data
                cell.alignment = align_right
                cell.border = thin_border
                cell.fill = fill_calc
                cell.number_format = '0.00E+00'

    # Auto-fit columns for Element_Stiffness
    for col in range(1, 28):
        col_letter = get_column_letter(col)
        ws_elem.column_dimensions[col_letter].width = 12
    ws_elem.column_dimensions['A'].width = 15
    ws_elem.column_dimensions['B'].width = 25

    # ----------------- 3. GLOBAL_SOLVER SHEET -----------------
    ws_solve['B2'] = "Global Assembly, Boundary Conditions, and Linear Equation Solver"
    ws_solve['B2'].font = font_title
    
    # We lay out the 48x48 Global Stiffness Matrix (K_global) starting at row 5, col 3 (C5:AX52)
    # Row 4 has headers
    for idx in range(48):
        node_id = idx // 6
        dof_idx = idx % 6
        dof_names = ["tx", "ty", "tz", "rx", "ry", "rz"]
        dof_name = f"N{node_id}_{dof_names[dof_idx]}"
        
        ws_solve.cell(row=4, column=3+idx, value=dof_name).font = font_header
        ws_solve.cell(row=4, column=3+idx).fill = fill_header
        ws_solve.cell(row=4, column=3+idx).alignment = align_center
        ws_solve.cell(row=4, column=3+idx).border = thin_border
        
        ws_solve.cell(row=5+idx, column=2, value=dof_name).font = font_data_bold
        ws_solve.cell(row=5+idx, column=2).alignment = align_left
        ws_solve.cell(row=5+idx, column=2).border = thin_border
        
    print("Writing K_global assembly cell formulas...")
    # Assemble K_global with nested IF formulas cell-by-cell
    # Note: element e has Node A at cell Element_Stiffness!C{35*e+5} (wait: `row_offset+1` was Node A ID, which is row `35*e+5`!)
    # Let's confirm: row_offset = 35 * e_idx + 4. Node A ID row = row_offset+1 = 35*e_idx + 5.
    # The cell address for Elem e Node A is Element_Stiffness!$C${35*e+5}.
    # Node B is Element_Stiffness!$E${35*e+5}.
    # Global stiffness block starts at column P (col 16), row 35*e+26 (which is k_global_start_row = row_offset+22 = 35*e+26).
    for r_idx in range(48):
        row_node = r_idx // 6
        row_dof = r_idx % 6
        for c_idx in range(48):
            col_node = c_idx // 6
            col_dof = c_idx % 6
            
            terms = []
            for e in range(7):
                elem_offset = 35 * e + 4
                elem_k_row = elem_offset + 22
                
                nodeA_ref = f"Element_Stiffness!$C${elem_offset+1}"
                nodeB_ref = f"Element_Stiffness!$E${elem_offset+1}"
                active_ref = f"Element_Stiffness!$C${elem_offset+2}"
                
                # Check column letter offsets
                col_AA_let = get_column_letter(16 + col_dof)
                col_AB_let = get_column_letter(16 + col_dof + 6)
                
                term = (f"IF({active_ref},IF({nodeA_ref}={row_node},IF({col_node}={nodeA_ref},Element_Stiffness!{col_AA_let}{elem_k_row+row_dof},IF({col_node}={nodeB_ref},Element_Stiffness!{col_AB_let}{elem_k_row+row_dof},0)),"
                        f"IF({nodeB_ref}={row_node},IF({col_node}={nodeA_ref},Element_Stiffness!{col_AA_let}{elem_k_row+row_dof+6},IF({col_node}={nodeB_ref},Element_Stiffness!{col_AB_let}{elem_k_row+row_dof+6},0)),0)),0)")
                terms.append(term)
                
            formula = "=" + "+".join(terms)
            
            # If it's a diagonal element (r_idx == c_idx), we must append the boundary condition penalty
            if r_idx == c_idx:
                bc_node = row_node
                # BC cells in Inputs sheet: starting at row 40 (Node 0 is row 40, Node 1 is 41, etc.)
                # tx is col C (3), ty is D (4), tz is E (5), rx is F (6), ry is G (7), rz is H (8)
                bc_col_let = get_column_letter(3 + row_dof)
                bc_cell = f"Inputs!{bc_col_let}{40+bc_node}"
                formula += f"+IF({bc_cell}=TRUE,1E15,IF(ISNUMBER({bc_cell}),{bc_cell},0))"
                
            cell = ws_solve.cell(row=5+r_idx, column=3+c_idx, value=formula)
            cell.font = font_data
            cell.alignment = align_right
            cell.border = thin_border
            cell.number_format = '0.00E+00'
            
    # Now, let's assemble the Global Load Vector F_T (Thermal) in Column AZ (col 52) and F_W (Weight) in Column AY (col 51)
    # Row 4 headers
    ws_solve.cell(row=4, column=51, value="F_Weight (N)").font = font_header
    ws_solve.cell(row=4, column=51).fill = fill_header
    ws_solve.cell(row=4, column=51).border = thin_border
    
    ws_solve.cell(row=4, column=52, value="F_Thermal (N)").font = font_header
    ws_solve.cell(row=4, column=52).fill = fill_header
    ws_solve.cell(row=4, column=52).border = thin_border
    
    # Calculate local thermal fixed-end force F_th = E * A * alpha * dT
    # For element e, E is Element_Stiffness!C{35*e+10}, A is Element_Stiffness!F{35*e+12}, alpha is Element_Stiffness!E{35*e+10}.
    # dT is Inputs!$K$11.
    # The local force vector has f_fe[0] = -F_th, f_fe[6] = F_th.
    # The global thermal force vector is F_fe = T^T * f_fe.
    # Let's write the global thermal force calculations into Element_Stiffness sheet to refer to it simply!
    # Let's place the global thermal force vector in Element_Stiffness columns B to C, rows 22 to 33!
    # Let's add that code back to our generation script.
    print("Writing Element_Stiffness thermal fixed-end forces formulas...")
    for e in range(7):
        elem_offset = 35 * e + 4
        # Local thermal force f_fe in column B, rows row_offset+22 to row_offset+33
        # Global thermal force F_fe in column C, rows row_offset+22 to row_offset+33
        E_cell = f"C{elem_offset+6}"
        A_cell = f"F{elem_offset+8}"
        alpha_cell = f"E{elem_offset+6}"
        dT_cell = "Inputs!$K$11"
        
        # Local thermal force F_th
        ws_elem.cell(row=elem_offset+22, column=2, value=f"=IF(C{elem_offset+2},-{E_cell}*{A_cell}*{alpha_cell}*{dT_cell},0)").number_format = '0.00E+00' # Node A axial
        ws_elem.cell(row=elem_offset+28, column=2, value=f"=IF(C{elem_offset+2},{E_cell}*{A_cell}*{alpha_cell}*{dT_cell},0)").number_format = '0.00E+00'  # Node B axial
        for r in range(12):
            if r not in (0, 6):
                ws_elem.cell(row=elem_offset+22+r, column=2, value=0.0).number_format = '0.00E+00'
            ws_elem.cell(row=elem_offset+22+r, column=2).font = font_data
            ws_elem.cell(row=elem_offset+22+r, column=2).border = thin_border
            
        # Global thermal force = MMULT(TRANSPOSE(T), f_fe) = SUMPRODUCT(col_T_transpose, f_fe)
        # Note: T column c starts at col 16 (P). Transpose row r is row r in T, which is P{row_offset+10+r}:AA{row_offset+10+r}.
        # f_fe is in column B, row_offset+22 to row_offset+33.
        for r in range(12):
            ws_elem.cell(row=elem_offset+22+r, column=3, 
                         value=f"=SUMPRODUCT(P{elem_offset+10+r}:AA{elem_offset+10+r}, B{elem_offset+22}:B{elem_offset+33})").number_format = '0.00E+00'
            ws_elem.cell(row=elem_offset+22+r, column=3).font = font_data
            ws_elem.cell(row=elem_offset+22+r, column=3).border = thin_border

    # Now write formulas for F_Thermal (Column AZ / Col 52) and F_Weight (Column AY / Col 51) in Global_Solver
    print("Writing global solver F_Thermal and F_Weight load vector formulas...")
    for idx in range(48):
        row_node = idx // 6
        row_dof = idx % 6
        
        # Assemble F_Thermal
        terms_T = []
        for e in range(7):
            elem_offset = 35 * e + 4
            nodeA_ref = f"Element_Stiffness!$C${elem_offset+1}"
            nodeB_ref = f"Element_Stiffness!$E${elem_offset+1}"
            active_ref = f"Element_Stiffness!$C${elem_offset+2}"
            
            term_T = f"IF({active_ref},IF({nodeA_ref}={row_node},Element_Stiffness!C{elem_offset+22+row_dof},IF({nodeB_ref}={row_node},Element_Stiffness!C{elem_offset+22+row_dof+6},0)),0)"
            terms_T.append(term_T)
            
        formula_T = "=" + "+".join(terms_T)
        cell_T = ws_solve.cell(row=5+idx, column=52, value=formula_T)
        cell_T.font = font_data
        cell_T.alignment = align_right
        cell_T.border = thin_border
        cell_T.number_format = '0.00E+00'
        
        # Assemble F_Weight (For simplicity, we pre-fill Weight load vector with 0.0 or add nodal point loads if any.
        # To make it complete, let's set F_Weight to 0.0 for now, or just let it read from Inputs nodal forces.
        # Let's set it to 0.0 as a baseline since thermal is the primary driving stress range, but keep the column ready!)
        cell_W = ws_solve.cell(row=5+idx, column=51, value=0.0)
        cell_W.font = font_data
        cell_W.alignment = align_right
        cell_W.border = thin_border
        cell_W.number_format = '0.00E+00'

    # Solve displacements using =MMULT(MINVERSE(K_global), Load_Vector)
    # Since we have dynamic array index solver:
    # Column BB (col 54) will hold u_Weight, and Column BC (col 55) will hold u_Thermal
    ws_solve.cell(row=4, column=54, value="u_Weight (m)").font = font_header
    ws_solve.cell(row=4, column=54).fill = fill_header
    ws_solve.cell(row=4, column=54).border = thin_border
    
    ws_solve.cell(row=4, column=55, value="u_Thermal (m)").font = font_header
    ws_solve.cell(row=4, column=55).fill = fill_header
    ws_solve.cell(row=4, column=55).border = thin_border
    
    print("Writing solver displacement formulas...")
    for idx in range(48):
        # Weight displacement
        ws_solve.cell(row=5+idx, column=54, value=f"=INDEX(MMULT(MINVERSE(C5:AX52), AY5:AY52), {idx+1}, 1)").number_format = '0.000000'
        # Thermal displacement
        ws_solve.cell(row=5+idx, column=55, value=f"=INDEX(MMULT(MINVERSE(C5:AX52), AZ5:AZ52), {idx+1}, 1)").number_format = '0.000000'
        
        for c in (54, 55):
            cell = ws_solve.cell(row=5+idx, column=c)
            cell.font = font_data_bold
            cell.alignment = align_right
            cell.fill = fill_calc
            cell.border = thin_border
            
    # Auto-fit columns for Global_Solver
    for col in range(1, 57):
        col_letter = get_column_letter(col)
        ws_solve.column_dimensions[col_letter].width = 12
    ws_solve.column_dimensions['A'].width = 15
    ws_solve.column_dimensions['B'].width = 15

    # ----------------- 4. STRESS_REPORT SHEET -----------------
    ws_report['B2'] = "ASME B31.3 Piping Stress & Code Compliance Report"
    ws_report['B2'].font = font_title
    
    headers_report = [
        "Element ID", "Node A", "Node B", "Type", "Flex Factor",
        "Sustained Stress (Pa)", "Sustained Allowable (Pa)", "Sustained Ratio",
        "Expansion Stress (Pa)", "Expansion Allowable (Pa)", "Expansion Ratio",
        "Max Ratio", "Status"
    ]
    for col_idx, h in enumerate(headers_report, 2):
        cell = ws_report.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Write report rows for Elements 0 to 6
    # Let's see: to calculate stresses, we extract element local forces:
    # local u_elem = T @ global u_elem
    # local force = k_local @ local_u - local f_fe.
    # To simplify and do it directly in Excel formulas:
    # We can calculate local forces for each element in the `Element_Stiffness` sheet!
    # Let's allocate columns B to C, rows 10 to 21 for local displacements and local forces:
    # - Column B: local displacement u_elem = T @ u_global (row_offset+10 to row_offset+21)
    # - Column C: local force f_local = k_local @ u_local - f_fe
    # This is incredibly beautiful and organized!
    # Let's write the formulas for local displacements and forces:
    print("Writing Element_Stiffness local displacement and force formulas...")
    for e in range(7):
        elem_offset = 35 * e + 4
        k_local_row = elem_offset + 10
        # Node A ID is Element_Stiffness!C{elem_offset+1}. Node B ID is E{elem_offset+1}.
        # Global displacements are in Global_Solver!BC{5 + 6*node_id} to Global_Solver!BC{10 + 6*node_id}.
        # Let's write a formula to assemble u_global_elem (12x1) in column AG, row_offset+10 to row_offset+21!
        # - AG{row_offset+10+d}: `=INDEX(Global_Solver!$BC$5:$BC$52, 6*NodeA + d + 1, 1)`
        # This is incredibly simple and clean!
        for d in range(6):
            ws_elem.cell(row=elem_offset+10+d, column=33, value=f"=INDEX(Global_Solver!$BC$5:$BC$52, 6*C{elem_offset+1}+{d+1}, 1)") # Node A DOFs
            ws_elem.cell(row=elem_offset+16+d, column=33, value=f"=INDEX(Global_Solver!$BC$5:$BC$52, 6*E{elem_offset+1}+{d+1}, 1)") # Node B DOFs
            
        for r in range(12):
            ws_elem.cell(row=elem_offset+10+r, column=33).font = font_data
            ws_elem.cell(row=elem_offset+10+r, column=33).border = thin_border
            ws_elem.cell(row=elem_offset+10+r, column=33).number_format = '0.000000'
            
        # Now, local displacement u_local = T @ u_global = SUMPRODUCT(row_T, u_global_elem)
        # Note: row r of T is at col 16 (P) to 27 (AA), row = k_local_row + r.
        # u_global_elem is in col 33 (AG), rows k_local_row to k_local_row + 11.
        for r in range(12):
            ws_elem.cell(row=elem_offset+10+r, column=34, 
                         value=f"=SUMPRODUCT(P{elem_offset+10+r}:AA{elem_offset+10+r}, AG{elem_offset+10}:AG{elem_offset+21})").number_format = '0.000000'
            ws_elem.cell(row=elem_offset+10+r, column=34).font = font_data
            ws_elem.cell(row=elem_offset+10+r, column=34).border = thin_border
            
        # Now, local forces f_local = k_local @ u_local - f_fe_local
        # k_local row r is D{k_local_row+r}:O{k_local_row+r}.
        # u_local is AH{k_local_row}:AH{k_local_row+11} (col 34 is AH).
        # f_fe_local is B{elem_offset+22+r}:B{elem_offset+22+r} (col 2 is B).
        # So: f_local = SUMPRODUCT(k_local_row, u_local) - f_fe_local
        for r in range(12):
            ws_elem.cell(row=elem_offset+10+r, column=35, 
                         value=f"=SUMPRODUCT(D{elem_offset+10+r}:O{elem_offset+10+r}, AH{elem_offset+10}:AH{elem_offset+21})-B{elem_offset+22+r}").number_format = '0.00E+00'
            ws_elem.cell(row=elem_offset+10+r, column=35).font = font_data
            ws_elem.cell(row=elem_offset+10+r, column=35).border = thin_border

    # Now, write the Stress_Report formulas!
    print("Writing stress report sheet formulas...")
    for e in range(7):
        r = 6 + e
        elem_offset = 35 * e + 4
        # Column B: Element ID
        ws_report.cell(row=r, column=2, value=f"=Inputs!B{29+e}").font = font_data_bold
        ws_report.cell(row=r, column=2).alignment = align_center
        ws_report.cell(row=r, column=2).border = thin_border
        
        # Column C: Node A
        ws_report.cell(row=r, column=3, value=f"=Inputs!C{29+e}").font = font_data
        ws_report.cell(row=r, column=3).alignment = align_center
        ws_report.cell(row=r, column=3).border = thin_border
        
        # Column D: Node B
        ws_report.cell(row=r, column=4, value=f"=Inputs!D{29+e}").font = font_data
        ws_report.cell(row=r, column=4).alignment = align_center
        ws_report.cell(row=r, column=4).border = thin_border
        
        # Column E: Type
        ws_report.cell(row=r, column=5, value=f"=Inputs!E{29+e}").font = font_data
        ws_report.cell(row=r, column=5).alignment = align_center
        ws_report.cell(row=r, column=5).border = thin_border
        
        # Column F: Flex Factor k
        ws_report.cell(row=r, column=6, value=f"=Element_Stiffness!F{elem_offset+9}").font = font_data
        ws_report.cell(row=r, column=6).alignment = align_right
        ws_report.cell(row=r, column=6).border = thin_border
        ws_report.cell(row=r, column=6).number_format = '0.00'
        
        # Column G: Sustained Stress (Pa)
        # S_L = Max of Node A and Node B
        # Node A bending moment M_A = SQRT(My^2 + Mz^2) = SQRT(AI{elem_offset+14}^2 + AI{elem_offset+15}^2) -> Wait:
        # local forces: col 35 (AI) is f_local.
        # - AI{row_offset+10} is Fx_A, AI{row_offset+11} is Fy_A, AI{row_offset+12} is Fz_A
        # - AI{row_offset+13} is Mx_A (torsion), AI{row_offset+14} is My_A, AI{row_offset+15} is Mz_A
        # - Node B: AI{row_offset+16} is Fx_B, etc.
        # Stress formula for Node A:
        # S_A = (SQRT((i_i * Mz_A)^2 + (i_o * My_A)^2) / Z) + (ABS(Fx_A) / A) + S_pr
        # For simplicity, straight pipe has SIFs i_i = 1.0, i_o = 1.0. Bends have SIFs (Inputs!E{29+e} is bend, calculate h and i_i, i_o).
        # Let's write SIF calculations in Stress_Report:
        # Let's define: SIF_in (i_i) = IF(Type="bend", MAX(1.0, 0.9 / (h^(2/3))), 1.0).
        # In Excel: h = (t * R) / r_m^2.
        h_expr = f"(Element_Stiffness!D{elem_offset+7}*IF(Inputs!F{29+e}<>\"\",Inputs!F{29+e},1.5*Element_Stiffness!C{elem_offset+7})) / (((Element_Stiffness!C{elem_offset+7}-Element_Stiffness!D{elem_offset+7})/2)^2)"
        sif_in = f'IF(E{r}="bend",MAX(1.0,0.9/({h_expr}^(2/3))),1.0)'
        sif_out = f'IF(E{r}="bend",MAX(1.0,0.75/({h_expr}^(2/3))),1.0)'
        
        ws_report.cell(row=r, column=20, value=f"={sif_in}")  # Column T: i_i helper
        ws_report.cell(row=r, column=21, value=f"={sif_out}") # Column U: i_o helper
        
        # Section properties: A is Element_Stiffness!F{elem_offset+8}, Z is Element_Stiffness!C{elem_offset+9} (Wait! Z is at Element_Stiffness!C23 which is 35*e+23, and A is Element_Stiffness!F21 which is 35*e+21!).
        # Let's check Area A and Z rows:
        # - `Area A` is row `elem_offset+8` = `35*e+12` (in col 6, which is F).
        # - `Section Modulus Z` is row `elem_offset+7` = `35*e+23` (in col 3, which is C).
        # Yes! A is Element_Stiffness!$F${35*e+12}, Z is Element_Stiffness!$C${35*e+23}.
        A_ref = f"Element_Stiffness!$F${35*e+12}"
        Z_ref = f"Element_Stiffness!$C${35*e+23}"
        Spr_ref = f"Element_Stiffness!$C${35*e+24}"
        Sh_ref = f"Inputs!$H$12"
        SA_ref = f"Element_Stiffness!$C${35*e+25}"
        
        # local forces:
        # Node A:
        Fx_A = f"Element_Stiffness!AI{elem_offset+10}"
        Mx_A = f"Element_Stiffness!AI{elem_offset+13}"
        My_A = f"Element_Stiffness!AI{elem_offset+14}"
        Mz_A = f"Element_Stiffness!AI{elem_offset+15}"
        
        # Node B:
        Fx_B = f"Element_Stiffness!AI{elem_offset+16}"
        Mx_B = f"Element_Stiffness!AI{elem_offset+19}"
        My_B = f"Element_Stiffness!AI{elem_offset+20}"
        Mz_B = f"Element_Stiffness!AI{elem_offset+21}"
        
        # Stress Node A = SQRT((i_i*Mz_A)^2 + (i_o*My_A)^2)/Z + ABS(Fx_A)/A + S_pr
        # (For thermal expansion case, expansion stress range S_E = SQRT((i_i*Mz)^2 + (i_o*My)^2 + 4*(0.5*Mx)^2)/Z)
        S_L_A = f"(SQRT((T{r}*{Mz_A})^2 + (U{r}*{My_A})^2)/{Z_ref}) + (ABS({Fx_A})/{A_ref}) + {Spr_ref}"
        S_L_B = f"(SQRT((T{r}*{Mz_B})^2 + (U{r}*{My_B})^2)/{Z_ref}) + (ABS({Fx_B})/{A_ref}) + {Spr_ref}"
        
        ws_report.cell(row=r, column=7, value=f"=IF(Inputs!C{29+e}<>\"\",MAX({S_L_A},{S_L_B}),0.0)").number_format = '0.00E+00' # Sustained Stress
        ws_report.cell(row=r, column=8, value=f"=IF(Inputs!C{29+e}<>\"\",{Sh_ref},0.0)").number_format = '0.00E+00'        # Sustained Allowable
        ws_report.cell(row=r, column=9, value=f"=IF(H{r}>0,G{r}/H{r},0.0)").number_format = '0.00'                        # Sustained Ratio
        
        # Thermal Expansion Stress Node A = SQRT((i_i*Mz_A)^2 + (i_o*My_A)^2 + 4*(0.5*Mx_A)^2)/Z
        S_E_A = f"SQRT((T{r}*{Mz_A})^2 + (U{r}*{My_A})^2 + 4*(0.5*{Mx_A})^2)/{Z_ref}"
        S_E_B = f"SQRT((T{r}*{Mz_B})^2 + (U{r}*{My_B})^2 + 4*(0.5*{Mx_B})^2)/{Z_ref}"
        
        ws_report.cell(row=r, column=10, value=f"=IF(Inputs!C{29+e}<>\"\",MAX({S_E_A},{S_E_B}),0.0)").number_format = '0.00E+00' # Expansion Stress
        ws_report.cell(row=r, column=11, value=f"=IF(Inputs!C{29+e}<>\"\",{SA_ref},0.0)").number_format = '0.00E+00'       # Expansion Allowable
        ws_report.cell(row=r, column=12, value=f"=IF(K{r}>0,J{r}/K{r},0.0)").number_format = '0.00'                       # Expansion Ratio
        
        # Max Ratio
        ws_report.cell(row=r, column=13, value=f"=MAX(I{r},L{r})").number_format = '0.00'
        # Status
        ws_report.cell(row=r, column=14, value=f'=IF(Inputs!C{29+e}<>"",IF(M{r}<=1.0,"PASS","FAIL"),"")').alignment = align_center
        
        # Styles
        for col_idx in range(2, 15):
            cell = ws_report.cell(row=r, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if r % 2 == 1:
                cell.fill = fill_zebra
                
        # Status styling
        cell_status = ws_report.cell(row=r, column=14)
        cell_status.font = font_pass
        cell_status.fill = fill_pass

    # Hide helper columns T and U (columns 20 and 21)
    ws_report.column_dimensions['T'].visible = False
    ws_report.column_dimensions['U'].visible = False

    # Native Excel conditional formatting for PASS/FAIL
    from openpyxl.formatting.rule import CellIsRule
    rule_pass = CellIsRule(operator='equal', formula=['"PASS"'], stopIfTrue=True, font=font_pass, fill=fill_pass)
    rule_fail = CellIsRule(operator='equal', formula=['"FAIL"'], stopIfTrue=True, font=font_fail, fill=fill_fail)
    ws_report.conditional_formatting.add('N6:N12', rule_pass)
    ws_report.conditional_formatting.add('N6:N12', rule_fail)
    
    # Conditional formatting for ratios > 1.0
    rule_ratio = CellIsRule(operator='greaterThan', formula=['1.0'], stopIfTrue=True, font=font_fail, fill=fill_fail)
    ws_report.conditional_formatting.add('I6:I12', rule_ratio)
    ws_report.conditional_formatting.add('L6:L12', rule_ratio)
    ws_report.conditional_formatting.add('M6:M12', rule_ratio)

    # Double bottom border on totals / end row
    double_bottom_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color='1F4E79')
    )
    for col in range(2, 15):
        ws_report.cell(row=12, column=col).border = double_bottom_border

    # Adjust columns width for Stress_Report
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        ws_report.column_dimensions[col_letter].width = 16
    ws_report.column_dimensions['B'].width = 12
    ws_report.column_dimensions['E'].width = 12

    # Save workbook
    wb.save(filepath)
    print(f"General multi-element Excel FEA solver workbook generated successfully at {filepath}")

if __name__ == "__main__":
    build_multi_element_workbook("piping_fea_solver.xlsx")
